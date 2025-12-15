import scrapy
from ..items import FragranceItem
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
import re

class MainSpiderScraper(scrapy.Spider):

    # Name for CLI call, syntax: scrapy crawl FragranceSpider
    name = "FragranceSpider"

    # URL from main page of website that will be scraped
    allowed_domains = ["kupujemprodajem.com", "www.kupujemprodajem.com"]

    # Specific URLs that you want to scrape from
    base_url = "https://www.kupujemprodajem.com/nega-i-licna-higijena/parfemi-muski/pretraga?categoryId=20&groupId=1314&keywords=ysl%20&page="

    # Define path where you want to save your xlsx file
    excel_path = r"C:\Users\lekid\OneDrive\Desktop\YSL.xlsx"

    # Number of pages to be scraped
    max_pages = 50

    #Conversion
    EUR_TO_RSD=117.0

    custom_settings = {
        'DOWNLOAD_DELAY': 1,
        'CONCURRENT_REQUESTS': 1,
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Fragrances"

        headers = ['Naslov', 'Cena (dinar)', 'Cena kontakt', 'Datum objave', 'Lokacija']
        self.worksheet.append(headers)

        #styling
        for cell in self.worksheet[1]:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

    def start_requests(self):
        yield scrapy.Request(
            url=f"{self.base_url}1",
            callback=self.parse,
            meta={'page': 1}
        )

    def convert_to_rsd(self, price_text):
        price_clean = price_text.replace(" ", "").replace(".", "")

        if "Kontakt" in price_clean:
            return None, "Kontakt"

        if "€" in price_clean or "eur" in price_clean.lower():
            match = re.search(r'(\d+)', price_clean)
            if match:
                eur_amount = int(match.group(1))
                rsd_amount = eur_amount * self.EUR_TO_RSD
                return rsd_amount, ""
        elif "din" in price_clean.lower() or "rsd" in price_clean.lower():
            match = re.search(r'(\d+)', price_clean)
            if match:
                rsd_amount = int(match.group(1))
                return rsd_amount, ""

        return None,"Kontakt"

    def parse(self, response):
        current_page = response.meta['page']

        fragrances = response.css(".AdItem_adOuterHolder__hb5N_")

        if not fragrances:
            print(f" No items found on page {current_page}. Stopping.")
            return

        items_scraped = 0

        # Data which will be scraped
        for i, fragrance in enumerate(fragrances):

            # Specific URL
            relative_url = fragrance.css("a::attr(href)").get()
            full_url = response.urljoin(relative_url)

            # Use this syntax to add more data if you need
            header = fragrance.css("div.AdItem_name__iOZvA::text").get()
            price = fragrance.css("div.AdItem_price__VZ_at div::text").get()

            price_rsd, original_price_text = self.convert_to_rsd(price) if self.convert_to_rsd(price) else (None, None)

            date = fragrance.xpath(".//p[svg and (contains(., 'dan') or contains(., 'juče') or contains(., 'pre dana'))]/text()[normalize-space()]").get()
            location = fragrance.css("div.AdItem_originAndPromoLocation__rQvKl p::text").get()

            # Only process if header exists
            if header:

                # Create Scrapy item
                item = FragranceItem()
                item["url"] = full_url
                item["header"] = header
                item["price_rsd"] = price_rsd
                item["date"] = date
                item["location"] = location
                item["price_contact"] = original_price_text
                yield item

                self.worksheet.append([
                header,
                price_rsd,
                original_price_text,
                date,
                location])

                current_row=self.worksheet.max_row

                if full_url:
                    link_cell = self.worksheet.cell(row=current_row, column=1)
                    link_cell.hyperlink = full_url

                    #fell free to change by taste
                    link_cell.font = Font(color="0563C1", underline="single")

                items_scraped += 1

        # Next page - only if items were found
        if items_scraped > 0 and current_page < self.max_pages:
            yield scrapy.Request(
                url=f"{self.base_url}{current_page + 1}",
                callback=self.parse,
                meta={'page': current_page + 1}
            )

    def closed(self, reason):
        directory = os.path.dirname(self.excel_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)

        self.workbook.save(self.excel_path)